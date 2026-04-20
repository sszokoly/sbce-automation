#!/usr/bin/python

from ansible.module_utils.basic import AnsibleModule
import openpyxl

def read_sbce_config(input_file):
    """Read excel and return data. Standardized for Ansible consumption."""
    sheets = []
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            keys = [cell.value for cell in ws['A']]
            values = [cell.value for cell in ws['B']]
            # Filters out empty rows to keep the fact clean
            sheets.append({sheet: {k: v for k, v in zip(keys, values) if k is not None}})
        return False, sheets, ""
    except Exception as e:
        return True, [], str(e)

def main():
    module = AnsibleModule(
        argument_spec=dict(
            path=dict(type='str', required=True),
        ),
        supports_check_mode=True
    )

    path = module.params['path']
    failed, result, error_msg = read_sbce_config(path)

    if failed:
        module.fail_json(msg="Failed to read Excel file: %s" % error_msg)

    # Returning as ansible_facts makes 'sbce_config' available to all following tasks
    module.exit_json(changed=False, ansible_facts={'sbce_config': result})

if __name__ == '__main__':
    main()