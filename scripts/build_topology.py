#!/usr/bin/env python3

from __future__ import annotations

import argparse
import ipaddress
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for index, ws in enumerate(workbook.worksheets):
        data = {}
        for row in ws.iter_rows(values_only=True):
            key = text(row[0] if row else None)
            if key:
                data[key] = row[1] if len(row) > 1 else None
        data["_sheet"] = ws.title
        data["_index"] = index
        sheets.append(data)
    return sheets


def parse_target_sheets(value: str | None) -> set[str] | None:
    if not value:
        return None
    result = {item.strip() for item in value.split(",") if item.strip()}
    return result or None


def ip_sort_key(value: Any) -> tuple[int, int | str]:
    raw = text(value)
    try:
        return (0, int(ipaddress.ip_address(raw)))
    except ValueError:
        return (1, raw)


def fail_if_empty(errors: list[str], sheet: dict[str, Any], name: str) -> str:
    value = text(sheet.get(name))
    if not value:
        errors.append(f"{sheet['_sheet']}: missing required field '{name}'")
    return value


def role_for_ems(node: dict[str, Any]) -> str:
    if text(node.get("ems_inst_type")).lower() == "secondary":
        return "EMS2"
    return "EMS1"


def main() -> int:
    parser = argparse.ArgumentParser(description="Build KVM test deployment topology from SBCE workbook")
    parser.add_argument("src", help="Workbook path")
    parser.add_argument("--target-sheets", default="", help="Comma-separated sheet names to execute exactly")
    args = parser.parse_args()

    path = Path(args.src)
    errors: list[str] = []
    try:
        all_sheets = load_workbook(path)
    except Exception as exc:
        print(json.dumps({"ok": False, "errors": [str(exc)]}))
        return 1

    selected_names = parse_target_sheets(args.target_sheets)
    all_names = {sheet["_sheet"] for sheet in all_sheets}
    if selected_names:
        missing = sorted(selected_names - all_names)
        for name in missing:
            errors.append(f"target_sheets: unknown sheet '{name}'")
        sheets = [sheet for sheet in all_sheets if sheet["_sheet"] in selected_names]
    else:
        sheets = all_sheets

    flat = {sheet["_sheet"]: sheet for sheet in sheets}

    platforms = {text(sheet.get("platform_type")).upper() for sheet in sheets if text(sheet.get("platform_type"))}
    if len(platforms) != 1:
        errors.append("selected sheets must contain exactly one platform_type")

    seen: dict[str, dict[str, str]] = {"ip0": {}, "vmname": {}, "hostname": {}, "appname": {}}
    for sheet in sheets:
        for field in ("platform_address", "platform_username", "platform_password", "ovf", "vmname", "hostname", "apptype", "ip0"):
            fail_if_empty(errors, sheet, field)
        for field in seen:
            value = text(sheet.get(field))
            if not value:
                continue
            previous = seen[field].get(value)
            if previous and not sheet.get("ha_peer_node"):
                errors.append(f"{sheet['_sheet']}: duplicate {field} '{value}' also used by {previous}")
            seen[field][value] = sheet["_sheet"]

    primary_by_ip: dict[str, dict[str, Any]] = {}
    primary_groups: dict[str, dict[str, Any]] = {}
    for sheet in sheets:
        apptype = text(sheet.get("apptype")).upper()
        ems_type = text(sheet.get("ems_inst_type")).lower() or "primary"
        if apptype in {"EMS", "EMS+SBCE"} and ems_type != "secondary":
            ip0 = text(sheet.get("ip0"))
            primary_by_ip[ip0] = sheet
            primary_groups[sheet["_sheet"]] = {
                "primary_sheet": sheet["_sheet"],
                "primary": dict(sheet, artifact_role="EMS1"),
                "coresident": apptype == "EMS+SBCE",
                "secondary_ems": [],
                "sbce_units": [],
            }

    if not primary_groups:
        errors.append("at least one primary EMS or EMS+SBCE sheet is required")

    child_by_ip = {text(sheet.get("ip0")): sheet for sheet in sheets if text(sheet.get("apptype")).upper() in {"SBCE", "EMS"}}
    unpaired: set[str] = set()

    for sheet in sheets:
        apptype = text(sheet.get("apptype")).upper()
        ems_type = text(sheet.get("ems_inst_type")).lower() or "primary"
        if apptype == "SBCE" or (apptype == "EMS" and ems_type == "secondary"):
            emsip = text(sheet.get("emsip"))
            primary = primary_by_ip.get(emsip)
            if not primary:
                errors.append(f"{sheet['_sheet']}: emsip '{emsip}' does not match a selected primary EMS ip0")
                continue
            group = primary_groups[primary["_sheet"]]
            if apptype == "EMS":
                group["secondary_ems"].append(dict(sheet, artifact_role="EMS2"))
            else:
                unpaired.add(sheet["_sheet"])

    for group in primary_groups.values():
        primary_ip = text(group["primary"].get("ip0"))
        sbces = [sheet for sheet in sheets if text(sheet.get("apptype")).upper() == "SBCE" and text(sheet.get("emsip")) == primary_ip]
        sbces.sort(key=lambda item: ip_sort_key(item.get("ip0")))
        for index, sheet in enumerate(sbces, start=1):
            sheet["artifact_role"] = f"SBC{index}"

        consumed: set[str] = set()
        for sheet in sbces:
            if sheet["_sheet"] in consumed:
                continue
            peer_ip = text(sheet.get("ha_peer_ip"))
            if peer_ip:
                peer = child_by_ip.get(peer_ip)
                if not peer or peer["_sheet"] not in flat:
                    errors.append(f"{sheet['_sheet']}: ha_peer_ip '{peer_ip}' does not match a selected SBCE ip0")
                    continue
                if text(peer.get("ha_peer_ip")) != text(sheet.get("ip0")):
                    errors.append(f"{sheet['_sheet']}: HA peer '{peer['_sheet']}' does not point back to this node")
                    continue
                if text(peer.get("emsip")) != primary_ip:
                    errors.append(f"{sheet['_sheet']}: HA peer '{peer['_sheet']}' belongs to a different EMS")
                    continue
                ordered = sorted([sheet, peer], key=lambda item: ip_sort_key(item.get("ip0")))
                group["sbce_units"].append({"type": "ha", "sheets": [node["_sheet"] for node in ordered], "nodes": [dict(node) for node in ordered]})
                consumed.update({sheet["_sheet"], peer["_sheet"]})
            else:
                group["sbce_units"].append({"type": "single", "sheets": [sheet["_sheet"]], "nodes": [dict(sheet)]})
                consumed.add(sheet["_sheet"])

    groups = list(primary_groups.values())
    groups.sort(key=lambda group: ip_sort_key(group["primary"].get("ip0")))

    if errors:
        print(json.dumps({"ok": False, "errors": errors}, indent=2))
        return 1

    print(json.dumps({"ok": True, "platform": "KVM", "groups": groups}, indent=2, default=str))
    return 0


if __name__ == "__main__":
    #sys.argv.extend(['data/xlsx/sbce_config_wrong.xlsx'])
    sys.exit(main())
