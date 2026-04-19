import openpyxl


def read_sbce_config(input_file):
    "Read the sbce_config.xlsx file and return content as Ansible facts"

    spreadsheet = {}
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            
            # Get cells from column A for keys
            column_A_cells = ws['A']
            keys = [cell.value for cell in column_A_cells]
            
            # Get cells from column (B) for values
            column_B_cells = ws['B']
            values = [cell.value for cell in column_B_cells]
            
            # Add sheet to 'sbce_config_sheets' dict
            spreadsheet[sheet] = {k:v for k,v, in zip(keys, values)}
    except IOError:
        return (1, "IOError on input file:%s" % input_file)

    result = {"ansible_facts": {"sbce_config_sheets": spreadsheet}}
    return (0, result)

if __name__ == "__main__":
    input_file =  "/home/sszokoly/Projects/sbce-automation/data/csv/sbce_config.xlsx"
    print(read_sbce_config(input_file))